import pandas as pd
import random
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Função para gerar o orçamento e salvar em uma planilha Excel
def gerar_orcamento():
    try:
        # Obtém o valor total desejado da interface gráfica
        valor_total = float(entry_valor.get())
        
        # Verifica se um arquivo CSV foi selecionado
        if not csv_path:
            messagebox.showerror("Erro", "Nenhum arquivo CSV selecionado.")
            return
        
        # Converte a tabela para um DataFrame do pandas
        df = pd.read_csv(csv_path, delimiter=";", skiprows=1, encoding='latin1')
        
        # Renomeia as colunas para facilitar o acesso
        df.columns = ['Descrição', 'Código', 'Unidade', 'Preço']
        
        # Remove possíveis espaços em branco nos nomes das colunas
        df.columns = df.columns.str.strip()
        
        # Converte a coluna "Preço" para numérico, tratando vírgulas como separadores decimais
        df['Preço'] = df['Preço'].str.replace('.', '').str.replace(',', '.').astype(float)
        
        # Verifica se há produtos no DataFrame
        if df.empty:
            messagebox.showerror("Erro", "Nenhum produto encontrado no arquivo CSV.")
            return
        
        # Lista para armazenar os produtos selecionados e suas quantidades
        produtos_selecionados = []
        valor_atual = 0
        
        # Define o valor máximo permitido (valor_total + margem de tolerância)
        margem_tolerancia = 0.1
        valor_maximo = valor_total * (1 + margem_tolerancia)
        
        # Enquanto o valor atual for menor que o valor máximo permitido
        while valor_atual < valor_maximo:
            # Escolhe um produto aleatoriamente
            produto = df.sample(1).iloc[0]
            
            # Escolhe uma quantidade aleatória para o produto (entre 1 e quantidade_maxima)
            quantidade_maxima = 10
            quantidade = random.randint(1, quantidade_maxima)
            
            # Calcula o valor total para esse produto
            valor_produto = produto['Preço'] * quantidade
            
            # Verifica se adicionar esse produto ultrapassa o valor máximo permitido
            if valor_atual + valor_produto > valor_maximo:
                # Ajusta a quantidade para não ultrapassar o valor máximo
                quantidade = int((valor_maximo - valor_atual) / produto['Preço'])
                if quantidade < 1:
                    break  # Se não for possível adicionar pelo menos 1 unidade, para o loop
            
            # Adiciona o produto e a quantidade à lista de selecionados
            produtos_selecionados.append({
                'Descrição': produto['Descrição'],
                'Código': produto['Código'],
                'Preço Unitário': produto['Preço'],
                'Quantidade': quantidade,
                'Valor Total': produto['Preço'] * quantidade
            })
            
            # Atualiza o valor atual
            valor_atual += produto['Preço'] * quantidade
        
        # Calcula o valor total dos produtos selecionados
        valor_total_produtos = sum(item['Valor Total'] for item in produtos_selecionados)
        
        # Calcula o desconto necessário (em valor)
        desconto = valor_total_produtos - valor_total
        if desconto < 0:
            desconto = 0  # Se o valor total dos produtos for menor que o desejado, não há desconto
        
        # Aplica o desconto
        valor_final = valor_total_produtos - desconto
        
        # Cria um arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Orçamento"
        
        # Define estilos para a planilha
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal="center", vertical="center")
        
        # Adiciona cabeçalhos à planilha
        ws.append(["Descrição", "Código", "Quantidade", "Preço Unitário", "Valor Total"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
        
        # Adiciona os produtos selecionados à planilha
        for item in produtos_selecionados:
            ws.append([item['Descrição'], item['Código'], item['Quantidade'], item['Preço Unitário'], item['Valor Total']])
        
        # Formata as células dos produtos
        for row in ws.iter_rows(min_row=2, max_row=len(produtos_selecionados) + 1, min_col=1, max_col=5):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
        
        # Adiciona o valor total, desconto e valor final à planilha
        ws.append([])  # Linha em branco
        ws.append(["Valor total dos produtos:", "", "", "", valor_total_produtos])
        ws.append(["Desconto aplicado:", "", "", "", desconto])
        ws.append(["Valor final do orçamento:", "", "", "", valor_final])
        
        # Formata as células dos valores totais
        for row in ws.iter_rows(min_row=len(produtos_selecionados) + 3, max_row=len(produtos_selecionados) + 5, min_col=1, max_col=5):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
                if cell.column == 5:  # Apenas a coluna de valores
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajusta o tamanho das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        
        # Salva o arquivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Orçamento salvo com sucesso em {file_path}")
    
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Função para selecionar o arquivo CSV
def selecionar_csv():
    global csv_path
    csv_path = filedialog.askopenfilename(filetypes=[("Arquivos CSV", "*.csv")])
    if csv_path:
        label_csv.config(text=f"Arquivo selecionado: {csv_path}")

# Cria a interface gráfica
root = tk.Tk()
root.title("Gerador de Orçamento")
root.geometry("500x250")

# Define um tema moderno
style = ttk.Style()
style.theme_use("clam")

# Frame para seleção do arquivo CSV
frame_csv = ttk.Frame(root)
frame_csv.pack(pady=10)

label_csv = ttk.Label(frame_csv, text="Nenhum arquivo CSV selecionado.")
label_csv.pack(side=tk.LEFT, padx=5)

button_csv = ttk.Button(frame_csv, text="Selecionar CSV", command=selecionar_csv)
button_csv.pack(side=tk.LEFT)

# Label e Entry para o valor total
label_valor = ttk.Label(root, text="Digite o valor total desejado para o orçamento: R$")
label_valor.pack(pady=10)

entry_valor = ttk.Entry(root)
entry_valor.pack(pady=5)

# Botão para gerar o orçamento
button_gerar = ttk.Button(root, text="Gerar Orçamento", command=gerar_orcamento)
button_gerar.pack(pady=20)

# Inicia o loop da interface gráfica
root.mainloop()
